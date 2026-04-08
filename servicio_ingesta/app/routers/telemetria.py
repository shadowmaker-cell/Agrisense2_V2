from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import List
from datetime import datetime, timezone
import httpx
import os
import io

from app.database import get_db
from app.models.lectura import LecturaSensor, LoteIngesta, ErrorIngesta, AlertaGenerada
from app.schemas.lectura import (
    LecturaEntrada, LoteEntrada,
    LecturaRespuesta, LoteRespuesta,
    AlertaRespuesta,
)
from app.services.validador import validar_lectura, detectar_alertas, normalizar_timestamp
from app.services.alertas import procesar_alertas
from app.events.producer import publish_telemetry_raw, publish_alert_generated, publish_batch_completed
from app.utils.jwt import get_usuario_id, get_usuario_id_opcional

router = APIRouter(prefix="/api/v1/telemetria", tags=["telemetria"])

_contadores = {}
LIMITE_POR_MINUTO  = 60
DEVICE_SERVICE_URL = os.getenv("DEVICE_SERVICE_URL", "http://localhost:8001")

VENTANAS = {
    "30m": "30 minutes",
    "1h":  "1 hour",
    "6h":  "6 hours",
    "24h": "24 hours",
    "7d":  "7 days",
}


# ── Utilidades ────────────────────────────────────────
def verificar_rate_limit(id_logico: str) -> bool:
    ahora         = datetime.now(timezone.utc)
    minuto_actual = ahora.strftime("%Y%m%d%H%M")
    clave         = f"{id_logico}:{minuto_actual}"
    _contadores[clave] = _contadores.get(clave, 0) + 1
    for k in [k for k in _contadores if not k.endswith(minuto_actual)]:
        del _contadores[k]
    return _contadores[clave] <= LIMITE_POR_MINUTO


def verificar_estado_dispositivo(id_logico: str) -> tuple:
    try:
        url      = f"{DEVICE_SERVICE_URL}/api/v1/dispositivos/"
        response = httpx.get(url, timeout=3.0)
        if response.status_code == 200:
            dispositivo = next(
                (d for d in response.json() if d.get("id_logico") == id_logico), None
            )
            if dispositivo and dispositivo.get("estado") != "activo":
                return False, f"Dispositivo {id_logico} en estado '{dispositivo['estado']}' — lecturas rechazadas"
    except Exception:
        pass
    return True, ""


def procesar_una_lectura(
    lectura_data: LecturaEntrada,
    db: Session,
    lote_id: int = None,
    usuario_id: int = None,
):
    if not verificar_rate_limit(lectura_data.id_logico):
        return None, "rate_limit", f"{lectura_data.id_logico} excedio {LIMITE_POR_MINUTO} lecturas/min"

    permitido, motivo = verificar_estado_dispositivo(lectura_data.id_logico)
    if not permitido:
        return None, "dispositivo_inactivo", motivo

    bandera, razon_error = validar_lectura(lectura_data.tipo_metrica, lectura_data.valor_metrica)
    ts = normalizar_timestamp(lectura_data.timestamp_lectura)

    lectura = LecturaSensor(
        usuario_id=usuario_id,
        dispositivo_id=lectura_data.dispositivo_id,
        id_logico=lectura_data.id_logico,
        tipo_metrica=lectura_data.tipo_metrica,
        valor_metrica=lectura_data.valor_metrica,
        unidad=lectura_data.unidad,
        timestamp_lectura=ts,
        bandera_calidad=bandera,
        lote_id=lote_id,
    )
    db.add(lectura)
    db.flush()
    publish_telemetry_raw(lectura)

    alertas_guardadas = procesar_alertas(
        db=db,
        dispositivo_id=lectura_data.dispositivo_id,
        id_logico=lectura_data.id_logico,
        tipo_metrica=lectura_data.tipo_metrica,
        valor=lectura_data.valor_metrica,
        alertas_detectadas=detectar_alertas(lectura_data.tipo_metrica, lectura_data.valor_metrica),
        usuario_id=usuario_id,
    )
    for alerta in alertas_guardadas:
        publish_alert_generated(alerta)

    return lectura, bandera, razon_error


# ── Lectura individual ────────────────────────────────
@router.post("/", response_model=LecturaRespuesta, status_code=201)
def recibir_lectura(
    payload: LecturaEntrada,
    request: Request,
    db: Session = Depends(get_db)
):
    usuario_id = get_usuario_id_opcional(request)
    lectura, bandera, razon_error = procesar_una_lectura(payload, db, usuario_id=usuario_id)

    if lectura is None:
        if bandera == "rate_limit":
            raise HTTPException(status_code=429, detail=razon_error)
        if bandera == "dispositivo_inactivo":
            raise HTTPException(status_code=403, detail=razon_error)
        raise HTTPException(status_code=400, detail=razon_error)

    if bandera == "invalido":
        db.add(ErrorIngesta(payload_raw=payload.model_dump_json(), razon_error=razon_error))

    db.commit()
    db.refresh(lectura)
    return lectura


# ── Lote ──────────────────────────────────────────────
@router.post("/lote", response_model=LoteRespuesta, status_code=201)
def recibir_lote(
    payload: LoteEntrada,
    request: Request,
    db: Session = Depends(get_db)
):
    usuario_id = get_usuario_id_opcional(request)
    lote = LoteIngesta(tipo_origen=payload.tipo_origen, total_registros=len(payload.lecturas))
    db.add(lote)
    db.flush()

    validos = invalidos = 0
    for lectura_data in payload.lecturas:
        lectura, bandera, razon_error = procesar_una_lectura(
            lectura_data, db, lote_id=lote.id, usuario_id=usuario_id
        )
        if lectura is None or bandera == "invalido":
            invalidos += 1
            db.add(ErrorIngesta(
                lote_id=lote.id,
                payload_raw=lectura_data.model_dump_json(),
                razon_error=razon_error,
            ))
        else:
            validos += 1

    lote.registros_validos   = validos
    lote.registros_invalidos = invalidos
    lote.estado = "procesado"
    db.commit()
    db.refresh(lote)
    publish_batch_completed(lote)

    return LoteRespuesta(
        lote_id=lote.id,
        total_registros=lote.total_registros,
        registros_validos=validos,
        registros_invalidos=invalidos,
        estado=lote.estado,
        alertas_generadas=0,
    )


# ── Ultimas lecturas ──────────────────────────────────
@router.get("/ultimas/{id_logico}", response_model=List[LecturaRespuesta])
def ultimas_lecturas(
    id_logico: str,
    request: Request,
    limite: int = 10,
    db: Session = Depends(get_db)
):
    usuario_id = get_usuario_id_opcional(request)
    query = db.query(LecturaSensor).filter(LecturaSensor.id_logico == id_logico)
    if usuario_id:
        query = query.filter(LecturaSensor.usuario_id == usuario_id)
    lecturas = query.order_by(LecturaSensor.timestamp_lectura.desc()).limit(limite).all()
    if not lecturas:
        raise HTTPException(status_code=404, detail=f"No se encontraron lecturas para {id_logico}")
    return lecturas


# ── Alertas ───────────────────────────────────────────
@router.get("/alertas", response_model=List[AlertaRespuesta])
def listar_alertas(
    request: Request,
    severidad: str = None,
    limite: int = 50,
    db: Session = Depends(get_db)
):
    usuario_id = get_usuario_id_opcional(request)
    query = db.query(AlertaGenerada)
    if usuario_id:
        query = query.filter(AlertaGenerada.usuario_id == usuario_id)
    if severidad:
        query = query.filter(AlertaGenerada.severidad == severidad)
    return query.order_by(AlertaGenerada.generada_en.desc()).limit(limite).all()


@router.get("/alertas/{id_logico}", response_model=List[AlertaRespuesta])
def alertas_por_dispositivo(
    id_logico: str,
    request: Request,
    limite: int = 20,
    db: Session = Depends(get_db)
):
    usuario_id = get_usuario_id_opcional(request)
    query = db.query(AlertaGenerada).filter(AlertaGenerada.id_logico == id_logico)
    if usuario_id:
        query = query.filter(AlertaGenerada.usuario_id == usuario_id)
    alertas = query.order_by(AlertaGenerada.generada_en.desc()).limit(limite).all()
    if not alertas:
        raise HTTPException(status_code=404, detail=f"No se encontraron alertas para {id_logico}")
    return alertas


# ── Promedios por ventana de tiempo (TimescaleDB) ─────
@router.get("/promedios/{id_logico}")
def promedios_sensor(
    id_logico: str,
    request: Request,
    ventana: str = "1h",
    db: Session = Depends(get_db)
):
    """
    Retorna promedios por ventana de tiempo.
    ventana: 30m | 1h | 6h | 24h | 7d
    """
    if ventana not in VENTANAS:
        raise HTTPException(
            status_code=400,
            detail=f"Ventana invalida. Opciones: {list(VENTANAS.keys())}"
        )

    usuario_id = get_usuario_id_opcional(request)
    intervalo  = VENTANAS[ventana]

    try:
        # TimescaleDB time_bucket
        sql = text("""
            SELECT
                time_bucket(:bucket, timestamp_lectura) AS periodo,
                tipo_metrica,
                ROUND(AVG(valor_metrica)::numeric, 2)   AS promedio,
                ROUND(MIN(valor_metrica)::numeric, 2)   AS minimo,
                ROUND(MAX(valor_metrica)::numeric, 2)   AS maximo,
                COUNT(*)                                AS total_lecturas,
                unidad
            FROM lectura_sensor
            WHERE id_logico  = :id_logico
              AND timestamp_lectura >= NOW() - :intervalo::interval
              AND (:usuario_id IS NULL OR usuario_id = :usuario_id)
            GROUP BY periodo, tipo_metrica, unidad
            ORDER BY periodo DESC, tipo_metrica
        """)
        rows = db.execute(sql, {
            "bucket":     intervalo,
            "id_logico":  id_logico,
            "intervalo":  intervalo,
            "usuario_id": usuario_id,
        }).fetchall()

    except Exception:
        # Fallback PostgreSQL estandar
        sql = text("""
            SELECT
                date_trunc('hour', timestamp_lectura)   AS periodo,
                tipo_metrica,
                ROUND(AVG(valor_metrica)::numeric, 2)   AS promedio,
                ROUND(MIN(valor_metrica)::numeric, 2)   AS minimo,
                ROUND(MAX(valor_metrica)::numeric, 2)   AS maximo,
                COUNT(*)                                AS total_lecturas,
                unidad
            FROM lectura_sensor
            WHERE id_logico  = :id_logico
              AND timestamp_lectura >= NOW() - :intervalo::interval
              AND (:usuario_id IS NULL OR usuario_id = :usuario_id)
            GROUP BY periodo, tipo_metrica, unidad
            ORDER BY periodo DESC, tipo_metrica
        """)
        rows = db.execute(sql, {
            "id_logico":  id_logico,
            "intervalo":  intervalo,
            "usuario_id": usuario_id,
        }).fetchall()

    if not rows:
        raise HTTPException(
            status_code=404,
            detail=f"No hay datos para {id_logico} en la ventana {ventana}"
        )

    return [
        {
            "periodo":        row.periodo.isoformat() if row.periodo else None,
            "tipo_metrica":   row.tipo_metrica,
            "promedio":       float(row.promedio)       if row.promedio else 0,
            "minimo":         float(row.minimo)         if row.minimo   else 0,
            "maximo":         float(row.maximo)         if row.maximo   else 0,
            "total_lecturas": row.total_lecturas,
            "unidad":         row.unidad,
        }
        for row in rows
    ]


# ── Export Excel ──────────────────────────────────────
@router.get("/export/{id_logico}/excel")
def exportar_excel(
    id_logico: str,
    request: Request,
    dias: int = 7,
    db: Session = Depends(get_db)
):
    """Exporta lecturas del sensor en Excel organizado por dia y hora con promedios."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    usuario_id = get_usuario_id_opcional(request)

    sql = text("""
        SELECT
            DATE(timestamp_lectura)                    AS fecha,
            EXTRACT(HOUR FROM timestamp_lectura)::int  AS hora,
            tipo_metrica,
            ROUND(AVG(valor_metrica)::numeric, 2)      AS promedio,
            ROUND(MIN(valor_metrica)::numeric, 2)      AS minimo,
            ROUND(MAX(valor_metrica)::numeric, 2)      AS maximo,
            COUNT(*)                                   AS total,
            unidad
        FROM lectura_sensor
        WHERE id_logico = :id_logico
          AND timestamp_lectura >= NOW() - :dias::interval
          AND (:usuario_id IS NULL OR usuario_id = :usuario_id)
        GROUP BY fecha, hora, tipo_metrica, unidad
        ORDER BY fecha DESC, hora DESC, tipo_metrica
    """)

    rows = db.execute(sql, {
        "id_logico":  id_logico,
        "dias":       f"{dias} days",
        "usuario_id": usuario_id,
    }).fetchall()

    if not rows:
        raise HTTPException(
            status_code=404,
            detail=f"No hay lecturas para {id_logico} en los ultimos {dias} dias"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Lecturas {id_logico}"

    # Estilos
    font_header  = Font(bold=True, color="FFFFFF")
    fill_header  = PatternFill("solid", fgColor="166534")
    fill_par     = PatternFill("solid", fgColor="f0fdf4")
    align_center = Alignment(horizontal="center")

    # Titulo
    ws.merge_cells("A1:H1")
    ws["A1"]           = f"Reporte de Lecturas — Sensor {id_logico}"
    ws["A1"].font      = Font(bold=True, size=14, color="166534")
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:H2")
    ws["A2"]           = f"Ultimos {dias} dias — AgriSense"
    ws["A2"].alignment = align_center

    # Headers
    headers = ["Fecha", "Hora", "Metrica", "Promedio", "Minimo", "Maximo", "Total Lecturas", "Unidad"]
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=4, column=col, value=h)
        cell.font      = font_header
        cell.fill      = fill_header
        cell.alignment = align_center

    # Datos
    for i, row in enumerate(rows, 5):
        valores = [
            str(row.fecha),
            f"{int(row.hora):02d}:00",
            row.tipo_metrica,
            float(row.promedio),
            float(row.minimo),
            float(row.maximo),
            int(row.total),
            row.unidad or "",
        ]
        for col, val in enumerate(valores, 1):
            cell           = ws.cell(row=i, column=col, value=val)
            cell.alignment = align_center
            if i % 2 == 0:
                cell.fill = fill_par

    # Anchos de columna
    for col, ancho in enumerate([14, 10, 22, 12, 12, 12, 16, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=lecturas_{id_logico}.xlsx"},
    )